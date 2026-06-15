#!/usr/bin/env python3
"""
Assemble the Springer Alt-Text deliverable from figures.json + alt_out/*.json.

Writes:
  accessibility/alt_text.xlsx  (the file to submit to Springer)
  accessibility/alt_text.csv   (UTF-8 with BOM; Excel opens natively)

The .xlsx is built directly from the stdlib (Office Open XML = zipped XML), so
no third-party packages are needed.
"""
import csv
import html
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
CHAPTERS = "/home/user/workspace/computational-etudes/textbook/chapters"
FIGS = os.path.join(HERE, "figures.json")
ALT_DIR = os.path.join(HERE, "alt_out")
XLSX = os.path.join(HERE, "alt_text.xlsx")
CSVF = os.path.join(HERE, "alt_text.csv")

HEADERS = ["Chapter", "Type", "Number", "Springer ID",
           "Source file", "Caption", "Alt Text"]
COL_WIDTHS = [26, 8, 12, 12, 42, 52, 80]


def load():
    recs = json.load(open(FIGS, encoding="utf-8"))
    alt = {}
    for fn in sorted(os.listdir(ALT_DIR)):
        if fn.endswith(".json"):
            d = json.load(open(os.path.join(ALT_DIR, fn), encoding="utf-8"))
            for k, v in d.items():
                if k in alt:
                    raise RuntimeError(f"duplicate springer_id across files: {k}")
                alt[k] = v
    return recs, alt


def chapter_titles(recs):
    titles = {}
    for r in recs:
        f = r["file"]
        if f in titles:
            continue
        txt = open(os.path.join(CHAPTERS, f), encoding="utf-8").read()
        m = re.search(r"(?m)^=\s+(.+)$", txt)
        t = m.group(1) if m else f
        t = re.sub(r"<[^>]+>", "", t)                 # strip <label>
        t = re.sub(r'#idx\(\s*"([^"]*)"\s*\)', r"\1", t)
        t = re.sub(r"#\w+(\([^)]*\))?", "", t)          # strip other macros
        titles[f] = t.strip()
    return titles


def source_file(r):
    if r["images"]:
        return "; ".join(p.replace("../", "") for p in r["images"])
    if r["kind"] == "table":
        return "live text table (no image file)"
    if r["kind"] == "drawn":
        return "drawn vector illustration (no image file)"
    return ""


def build_rows(recs, alt, titles):
    rows = []
    for r in recs:
        ch = r["chapter"]
        chap = f"Chapter {ch}: {titles[r['file']]}" if isinstance(ch, int) \
            else titles.get(r["file"], str(ch))
        typ = "Table" if r["kind"] == "table" else "Figure"
        rows.append([
            chap,
            typ,
            r["number"],
            r["springer_id"],
            source_file(r),
            r["caption"],
            alt.get(r["springer_id"], ""),
        ])
    return rows


# ---------- minimal XLSX writer ----------
def col_letter(n):
    s = ""
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def clean_xml(s):
    s = s.replace("\r", " ").replace("\n", " ")
    s = "".join(c for c in s if ord(c) >= 0x20 or c == "\t")
    return html.escape(s, quote=False)


def cell_xml(col, row, text, style):
    ref = f"{col_letter(col)}{row}"
    return (f'<c r="{ref}" s="{style}" t="inlineStr">'
            f'<is><t xml:space="preserve">{clean_xml(text)}</t></is></c>')


def sheet_xml(header, rows):
    out = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
           '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
           '<sheetViews><sheetView workbookViewId="0">'
           '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
           '</sheetView></sheetViews>',
           '<cols>']
    for i, w in enumerate(COL_WIDTHS, 1):
        out.append(f'<col min="{i}" max="{i}" width="{w}" customWidth="1"/>')
    out.append("</cols><sheetData>")
    out.append('<row r="1">')
    for c, h in enumerate(header, 1):
        out.append(cell_xml(c, 1, h, 1))
    out.append("</row>")
    for ri, row in enumerate(rows, 2):
        out.append(f'<row r="{ri}">')
        for c, val in enumerate(row, 1):
            out.append(cell_xml(c, ri, str(val), 2))
        out.append("</row>")
    out.append("</sheetData></worksheet>")
    return "".join(out)


STYLES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font>
<font><b/><sz val="11"/><name val="Calibri"/></font></fonts>
<fills count="2"><fill><patternFill patternType="none"/></fill>
<fill><patternFill patternType="gray125"/></fill></fills>
<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="3">
<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1" applyAlignment="1"><alignment horizontal="left" vertical="top" wrapText="1"/></xf>
<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
</cellXfs>
<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""

CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""

RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""

WORKBOOK = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="Alt Text" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""

WORKBOOK_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""


def write_xlsx_openpyxl(path, header, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Alt Text"
    ws.append(header)
    for row in rows:
        ws.append([str(c) for c in row])
    bold = Font(bold=True)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    head_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).alignment = head_align
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c - 1]
    for r in range(2, len(rows) + 2):
        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).alignment = wrap_top
    ws.freeze_panes = "A2"
    wb.save(path)


def validate_xlsx_openpyxl(path, n_rows):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == n_rows + 1, f"row count {ws.max_row} != {n_rows + 1}"
    return ws.max_row


def write_xlsx(path, header, rows):
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", CONTENT_TYPES)
        z.writestr("_rels/.rels", RELS)
        z.writestr("xl/workbook.xml", WORKBOOK)
        z.writestr("xl/_rels/workbook.xml.rels", WORKBOOK_RELS)
        z.writestr("xl/styles.xml", STYLES)
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml(header, rows))


def validate_xlsx(path, n_rows):
    """Re-open and parse to confirm the workbook is well-formed."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        for req in ["[Content_Types].xml", "xl/workbook.xml",
                    "xl/worksheets/sheet1.xml", "xl/styles.xml"]:
            assert req in names, f"missing part {req}"
        data = z.read("xl/worksheets/sheet1.xml")
    root = ET.fromstring(data)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rows = root.findall(f".//{ns}row")
    assert len(rows) == n_rows + 1, f"row count {len(rows)} != {n_rows + 1}"
    return len(rows)


def main():
    recs, alt = load()
    titles = chapter_titles(recs)

    known = {r["springer_id"] for r in recs}
    missing = [r["springer_id"] for r in recs if r["springer_id"] not in alt]
    extra = [k for k in alt if k not in known]
    if missing:
        print(f"WARNING: {len(missing)} figures missing alt text: {missing}")
    if extra:
        print(f"WARNING: {len(extra)} alt-text keys not in figure list: {extra}")

    rows = build_rows(recs, alt, titles)

    # CSV (UTF-8 with BOM, every field quoted)
    with open(CSVF, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(HEADERS)
        w.writerows(rows)

    # XLSX (prefer openpyxl; fall back to stdlib writer)
    try:
        import openpyxl  # noqa: F401
        write_xlsx_openpyxl(XLSX, HEADERS, rows)
        nrows = validate_xlsx_openpyxl(XLSX, len(rows))
        writer = "openpyxl"
    except ImportError:
        write_xlsx(XLSX, HEADERS, rows)
        nrows = validate_xlsx(XLSX, len(rows))
        writer = "stdlib"

    n_fig = sum(1 for r in rows if r[1] == "Figure")
    n_tbl = sum(1 for r in rows if r[1] == "Table")
    print(f"OK: {len(rows)} rows ({n_fig} figures, {n_tbl} tables)")
    print(f"  xlsx: {XLSX}  ({nrows} rows incl. header, validated via {writer})")
    print(f"  csv : {CSVF}")
    if not missing and not extra:
        print("  alt-text coverage: 100% (no missing, no extras)")


if __name__ == "__main__":
    main()
